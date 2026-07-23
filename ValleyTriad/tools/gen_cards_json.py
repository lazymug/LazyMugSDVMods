#!/usr/bin/env python3
"""Reads cards.xlsx (single source) and emits assets/cards.json + merges card.* i18n keys."""
import json, re, collections
import openpyxl

ROOT = "/Volumes/SSD_Work/Workspace/LazyMugSDVMods/ValleyTriad"
XLSX = f"{ROOT}/cards.xlsx"
OUT_JSON = f"{ROOT}/assets/cards.json"

TIER = {"Comum": "Common", "Incomum": "Uncommon", "Raro": "Rare", "Lendário": "Legendary"}
ELEM = {"Primavera": "Spring", "Verão": "Summer", "Outono": "Fall", "Inverno": "Winter", "Nenhum": "None"}
CAT = {"Crop": "Crop", "Forrageável": "Forage", "Animal": "Animal", "Peixe": "Fish",
       "Monstro": "Monster", "Mineral": "Mineral", "Aldeão": "Villager", "Especial": "Special"}

# remap prototype sprite refs -> real game asset names
SPRITE = {
    "Animal:Chicken": "Animal:White Chicken", "Animal:Cow": "Animal:White Cow",
    "Animal:Duck": "Animal:Duck", "Animal:Rabbit": "Animal:Rabbit",
    "Monster:Skeleton": "Monster:Skeleton", "Monster:ShadowBrute": "Monster:Shadow Brute",
    "Monster:Serpent": "Monster:Serpent", "Monster:PepperRex": "Monster:Pepper Rex",
    "Villager:Qi": "Villager:MrQi", "Villager:Wizard": "Villager:Wizard",
    "Special:JunimoKing": "Char:Junimo",
}

def slug(name): return re.sub(r"[^a-z0-9]", "", name.lower())

wb = openpyxl.load_workbook(XLSX)
ws = wb["Cartas"]
cards, en, pt = [], {}, {}
for r in range(5, 70):
    en_name = ws.cell(r, 2).value
    tier = ws.cell(r, 5).value
    if not en_name or not tier:
        continue
    pt_name = ws.cell(r, 3).value or en_name
    cat = ws.cell(r, 4).value
    elem = ws.cell(r, 6).value
    sprite = ws.cell(r, 7).value
    n, e, s, w = (ws.cell(r, c).value for c in (8, 9, 10, 11))
    sidx = ws.cell(r, 16).value or 0
    if None in (cat, elem, sprite, n, e, s, w):
        continue
    cid = slug(en_name)
    key = f"card.{cid}"
    cards.append({
        "Id": cid, "NameKey": key,
        "Tier": TIER[tier], "Element": ELEM[elem], "Category": CAT[cat],
        "Sprite": SPRITE.get(sprite, sprite), "SpriteIndex": int(sidx),
        "N": int(n), "E": int(e), "S": int(s), "W": int(w),
    })
    en[key] = en_name
    pt[key] = pt_name

with open(OUT_JSON, "w", encoding="utf-8") as f:
    json.dump(cards, f, ensure_ascii=False, indent=2)

def merge_i18n(path, names):
    with open(path, encoding="utf-8") as f:
        data = json.load(f, object_pairs_hook=collections.OrderedDict)
    data = collections.OrderedDict((k, v) for k, v in data.items() if not k.startswith("card."))
    for k, v in names.items():
        data[k] = v
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

merge_i18n(f"{ROOT}/i18n/default.json", en)
merge_i18n(f"{ROOT}/i18n/pt-BR.json", pt)
print(f"cards.json: {len(cards)} cartas · i18n atualizado")
