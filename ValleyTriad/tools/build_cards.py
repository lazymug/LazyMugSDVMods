#!/usr/bin/env python3
"""Gera ValleyTriad/cards.xlsx — acervo do Valley Triad, preenchido com objetos reais do SDV."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Volumes/SSD_Work/Workspace/LazyMugSDVMods/ValleyTriad/cards.xlsx"

wb = openpyxl.Workbook()

HDR_FILL = PatternFill("solid", fgColor="4472C4")
HDR_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TIER_FILL = {
    "Comum":    PatternFill("solid", fgColor="E2EFDA"),
    "Incomum":  PatternFill("solid", fgColor="DDEBF7"),
    "Raro":     PatternFill("solid", fgColor="FCE4D6"),
    "Lendário": PatternFill("solid", fgColor="FFF2CC"),
}

# ---------- aba Tiers ----------
wt = wb.active
wt.title = "Tiers"
wt["A1"] = "Orçamento de borda por tier"; wt["A1"].font = TITLE_FONT
for c, h in enumerate(["Tier", "Soma mín", "Soma máx", "Qtd alvo (v1)", "Observação"], 1):
    cell = wt.cell(3, c, h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
tiers = [
    ("Comum",    8,  15, 24, "crops, forrageáveis, animais"),
    ("Incomum",  16, 22, 15, "aldeões"),
    ("Raro",     23, 29, 8,  "monstros, minerais, peixes/goods especiais"),
    ("Lendário", 30, 36, 5,  "Mago, Rei Junimo, Sr. Qi, peixe lendário, Frag. Prismático"),
]
for r, row in enumerate(tiers, 4):
    for c, v in enumerate(row, 1):
        cell = wt.cell(r, c, v); cell.border = BORDER
        cell.alignment = CENTER if c in (2, 3, 4) else LEFT
        if c == 1:
            cell.fill = TIER_FILL[row[0]]; cell.font = Font(bold=True)
wt.cell(9, 1, "Cada carta tem 4 bordas (N/L/S/O), 1 a 10 (10 = 'A'). A soma das 4 deve cair na faixa do tier.")
for col, w in zip("ABCDE", (12, 10, 10, 14, 46)):
    wt.column_dimensions[col].width = w

# ---------- aba Cartas ----------
# Dados extraídos do Content/Data/Objects.xnb do jogo (v1.6.15): id -> (SpriteIndex, Preço)
GAME_DATA = {
    16:(16,50),24:(24,35),72:(72,750),74:(74,2000),80:(80,25),159:(159,1500),188:(188,40),
    190:(190,175),192:(192,80),254:(254,250),256:(256,60),258:(258,50),260:(260,40),272:(272,60),
    276:(276,320),280:(280,160),282:(282,75),296:(296,5),396:(396,80),404:(404,40),410:(410,20),
    412:(412,70),416:(416,100),417:(417,3000),454:(454,550),698:(698,200),
}

ws = wb.create_sheet("Cartas")
cols = [
    ("ID", 6), ("Nome (EN)", 20), ("Nome (PT)", 22), ("Categoria", 13), ("Tier", 11),
    ("Elemento (estação)", 17), ("Sprite (fonte/ID)", 20),
    ("N", 5), ("L", 5), ("S", 5), ("O", 5), ("Soma", 7), ("Orçamento?", 11),
    ("Aquisição", 15), ("Notas", 26), ("SpriteIndex", 12), ("Preço (g)", 10),
]
ws["A1"] = "Valley Triad — Acervo de cartas (v1)"; ws["A1"].font = TITLE_FONT
ws["A2"] = ("Objetos reais do SDV (wiki). Sprite = de onde puxar a imagem em runtime: "
            "(O)<id> objeto · Villager:<Nome> · Animal:<Tipo> · Monster:<Tipo> · Special:<Nome>. "
            "'Soma' e 'Orçamento?' são automáticos.")
ws["A2"].font = Font(italic=True, color="808080")

HROW = 4
for c, (h, w) in enumerate(cols, 1):
    cell = ws.cell(HROW, c, h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "A5"

# (ID, EN, PT, Categoria, Tier, Elemento, Sprite, N, L, S, O, Aquisição, Notas)
# Balanceamento: força escala dentro do tier; cada carta tem um "formato" (pico numa
# direção, fraqueza na oposta) para dar leitura tática. IDs validados no Objects.xnb (1.6.15).
seed = [
    # ---- Comum: crops ----
    (1,  "Parsnip",         "Pastinaga",           "Crop",        "Comum",   "Primavera", "(O)24",  3, 1, 2, 2, "Pacote inicial", "starter fraco"),
    (2,  "Cauliflower",     "Couve-flor",          "Crop",        "Comum",   "Primavera", "(O)190", 5, 4, 3, 2, "Pacote inicial", "crop grande"),
    (3,  "Potato",          "Batata",              "Crop",        "Comum",   "Primavera", "(O)192", 3, 3, 3, 2, "Pacote inicial", "arredondada"),
    (4,  "Green Bean",      "Vagem",               "Crop",        "Comum",   "Primavera", "(O)188", 2, 5, 2, 2, "Ganhar de NPC", "pico à direita"),
    (5,  "Blueberry",       "Mirtilo",             "Crop",        "Comum",   "Verão",     "(O)258", 4, 2, 4, 3, "Pacote inicial", "vertical"),
    (6,  "Hot Pepper",      "Pimenta",             "Crop",        "Comum",   "Verão",     "(O)260", 5, 2, 3, 1, "Ganhar de NPC", "pico no topo"),
    (7,  "Melon",           "Melão",               "Crop",        "Comum",   "Verão",     "(O)254", 4, 4, 4, 2, "Ganhar de NPC", "fruta grande"),
    (8,  "Tomato",          "Tomate",              "Crop",        "Comum",   "Verão",     "(O)256", 3, 4, 3, 2, "Pacote inicial", ""),
    (9,  "Pumpkin",         "Abóbora",             "Crop",        "Comum",   "Outono",    "(O)276", 6, 3, 4, 1, "Ganhar de NPC", "topo forte; casa de Outono"),
    (10, "Cranberries",     "Oxicoco",             "Crop",        "Comum",   "Outono",    "(O)282", 3, 4, 3, 4, "Ganhar de NPC", "horizontal"),
    (11, "Eggplant",        "Berinjela",           "Crop",        "Comum",   "Outono",    "(O)272", 4, 3, 4, 2, "Ganhar de NPC", ""),
    (12, "Yam",             "Inhame",              "Crop",        "Comum",   "Outono",    "(O)280", 3, 2, 5, 2, "Pacote inicial", "raiz: pico embaixo"),
    # ---- Comum: forrageáveis (mais fracos) ----
    (13, "Wild Horseradish","Raiz-forte silvestre","Forrageável", "Comum",   "Primavera", "(O)16",  2, 3, 2, 1, "Ganhar de NPC", ""),
    (14, "Salmonberry",     "Amora-salmão",        "Forrageável", "Comum",   "Primavera", "(O)296", 2, 2, 3, 2, "Ganhar de NPC", ""),
    (15, "Spice Berry",     "Baga Picante",        "Forrageável", "Comum",   "Verão",     "(O)396", 3, 2, 2, 2, "Ganhar de NPC", ""),
    (16, "Common Mushroom", "Cogumelo Comum",      "Forrageável", "Comum",   "Outono",    "(O)404", 3, 3, 2, 2, "Ganhar de NPC", ""),
    (17, "Blackberry",      "Amora-preta",         "Forrageável", "Comum",   "Outono",    "(O)410", 2, 3, 3, 2, "Ganhar de NPC", ""),
    (18, "Winter Root",     "Raiz de Inverno",     "Forrageável", "Comum",   "Inverno",   "(O)412", 3, 2, 3, 2, "Ganhar de NPC", "raro no Inverno"),
    (19, "Snow Yam",        "Inhame da Neve",      "Forrageável", "Comum",   "Inverno",   "(O)416", 2, 3, 2, 3, "Ganhar de NPC", ""),
    # ---- Comum: animais + mineral básico ----
    (20, "Chicken",         "Galinha",             "Animal",      "Comum",   "Nenhum",    "Animal:Chicken", 2, 2, 3, 2, "Pacote inicial", ""),
    (21, "Cow",             "Vaca",                "Animal",      "Comum",   "Nenhum",    "Animal:Cow",     4, 3, 4, 3, "Ganhar de NPC", "animal grande"),
    (22, "Duck",            "Pato",                "Animal",      "Comum",   "Nenhum",    "Animal:Duck",    3, 3, 2, 3, "Ganhar de NPC", ""),
    (23, "Rabbit",          "Coelho",              "Animal",      "Comum",   "Nenhum",    "Animal:Rabbit",  2, 4, 2, 3, "Ganhar de NPC", "ágil: laterais"),
    (24, "Quartz",          "Quartzo",             "Mineral",     "Comum",   "Nenhum",    "(O)80",  4, 2, 2, 4, "Ganhar de NPC", "cristal: horizontal"),
    # ---- Incomum: aldeões ----
    (25, "Abigail",   "Abigail",   "Aldeão", "Incomum", "Nenhum", "Villager:Abigail",   6, 6, 4, 5, "Ganhar de NPC", "mentora / evento inicial"),
    (26, "Sebastian", "Sebastian", "Aldeão", "Incomum", "Nenhum", "Villager:Sebastian", 5, 6, 5, 5, "Ganhar de NPC", ""),
    (27, "Sam",       "Sam",       "Aldeão", "Incomum", "Nenhum", "Villager:Sam",       5, 5, 4, 4, "Ganhar de NPC", ""),
    (28, "Penny",     "Penny",     "Aldeão", "Incomum", "Nenhum", "Villager:Penny",     4, 4, 6, 4, "Ganhar de NPC", "acolhedora: base forte"),
    (29, "Leah",      "Leah",      "Aldeão", "Incomum", "Nenhum", "Villager:Leah",      5, 4, 5, 4, "Ganhar de NPC", ""),
    (30, "Elliott",   "Elliott",   "Aldeão", "Incomum", "Nenhum", "Villager:Elliott",   4, 6, 4, 5, "Ganhar de NPC", "eloquente: direita"),
    (31, "Haley",     "Haley",     "Aldeão", "Incomum", "Nenhum", "Villager:Haley",     5, 4, 4, 4, "Ganhar de NPC", ""),
    (32, "Maru",      "Maru",      "Aldeão", "Incomum", "Nenhum", "Villager:Maru",      5, 5, 5, 5, "Ganhar de NPC", "engenheira: equilibrada"),
    (33, "Emily",     "Emily",     "Aldeão", "Incomum", "Nenhum", "Villager:Emily",     4, 5, 4, 5, "Ganhar de NPC", ""),
    (34, "Shane",     "Shane",     "Aldeão", "Incomum", "Nenhum", "Villager:Shane",     6, 4, 5, 3, "Ganhar de NPC", "pico no topo"),
    (35, "Gus",       "Gus",       "Aldeão", "Incomum", "Nenhum", "Villager:Gus",       5, 5, 6, 5, "Ganhar de NPC", "dono do Saloon; robusto"),
    (36, "Pierre",    "Pierre",    "Aldeão", "Incomum", "Nenhum", "Villager:Pierre",    5, 4, 4, 4, "Ganhar de NPC", "mediano"),
    (37, "Robin",     "Robin",     "Aldeão", "Incomum", "Nenhum", "Villager:Robin",     6, 5, 5, 4, "Ganhar de NPC", "carpinteira: sólida"),
    (38, "Clint",     "Clint",     "Aldeão", "Incomum", "Nenhum", "Villager:Clint",     6, 5, 4, 5, "Ganhar de NPC", "ferreiro: forte"),
    (39, "Willy",     "Willy",     "Aldeão", "Incomum", "Nenhum", "Villager:Willy",     4, 5, 4, 5, "Ganhar de NPC", "pescador"),
    # ---- Raro: monstros, minerais, peixes/goods ----
    (40, "Skeleton",     "Esqueleto",        "Monstro", "Raro", "Nenhum", "Monster:Skeleton",   7, 6, 6, 5, "Ganhar de NPC", ""),
    (41, "Shadow Brute", "Bruto das Sombras","Monstro", "Raro", "Nenhum", "Monster:ShadowBrute",8, 6, 6, 5, "Ganhar de NPC", "agressivo: topo"),
    (42, "Serpent",      "Serpente",         "Monstro", "Raro", "Nenhum", "Monster:Serpent",    6, 8, 6, 4, "Ganhar de NPC", "veloz: direita"),
    (43, "Pepper Rex",   "Pepper Rex",       "Monstro", "Raro", "Nenhum", "Monster:PepperRex",  8, 7, 7, 5, "Ganhar de NPC", "dino poderoso"),
    (44, "Diamond",      "Diamante",         "Mineral", "Raro", "Nenhum", "(O)72",   7, 6, 7, 6, "Ganhar de NPC", "equilíbrio alto"),
    (45, "Ancient Fruit","Fruta Ancestral",  "Crop",    "Raro", "Nenhum", "(O)454",  7, 7, 6, 6, "Ganhar de NPC", ""),
    (46, "Sturgeon",     "Esturjão",         "Peixe",   "Raro", "Nenhum", "(O)698",  6, 6, 8, 5, "Ganhar de NPC", "peixe grande: base"),
    (47, "Sweet Gem Berry","Baga-joia Doce", "Crop",    "Raro", "Nenhum", "(O)417",  8, 7, 7, 6, "Ganhar de NPC", "premium do tier"),
    # ---- Lendário: chefes + peixe lendário + relíquia ----
    (48, "Wizard",          "Mago",               "Aldeão",   "Lendário", "Nenhum", "Villager:Wizard",     9, 8, 7, 8, "Campeão", "boss"),
    (49, "Junimo King",     "Rei Junimo",         "Especial", "Lendário", "Nenhum", "Special:JunimoKing",  8, 9, 8, 9, "Campeão", "boss"),
    (50, "Mr. Qi",          "Sr. Qi",             "Especial", "Lendário", "Nenhum", "Villager:Qi",        10, 9, 8, 8, "Campeão", "boss supremo"),
    (51, "Crimsonfish",     "Peixe Carmesim",     "Peixe",    "Lendário", "Nenhum", "(O)159",              9, 8, 9, 7, "Campeão", "peixe lendário"),
    (52, "Prismatic Shard", "Fragmento Prismático","Mineral", "Lendário", "Nenhum", "(O)74",              10, 9, 9, 8, "Campeão", "relíquia máxima"),
]

TOTAL_ROWS = 58  # 52 preenchidas + linhas em branco para expandir
first = HROW + 1
for i in range(TOTAL_ROWS):
    r = first + i
    data = seed[i] if i < len(seed) else None
    vals = list(data[:11]) if data else [""] * 11
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.border = BORDER
        cell.alignment = CENTER if c in (1, 5, 6, 8, 9, 10, 11) else LEFT
        if c == 5 and v in TIER_FILL:
            cell.fill = TIER_FILL[v]
    som = ws.cell(r, 12); som.border = BORDER; som.alignment = CENTER
    som.value = f'=IF(COUNT(H{r}:K{r})=4,SUM(H{r}:K{r}),"")'
    ok = ws.cell(r, 13); ok.border = BORDER; ok.alignment = CENTER
    ok.value = (
        f'=IF(OR(E{r}="",L{r}=""),"",'
        f'IF(AND(L{r}>=VLOOKUP(E{r},Tiers!$A$4:$C$7,2,0),'
        f'L{r}<=VLOOKUP(E{r},Tiers!$A$4:$C$7,3,0)),"OK","Fora"))'
    )
    for c in (14, 15):
        cell = ws.cell(r, c); cell.border = BORDER; cell.alignment = LEFT
    # colunas de dados do jogo (SpriteIndex, Preço) — só para cartas de objeto (O)id
    si = ws.cell(r, 16); si.border = BORDER; si.alignment = CENTER
    pr = ws.cell(r, 17); pr.border = BORDER; pr.alignment = CENTER
    if data:
        ws.cell(r, 14, data[11]); ws.cell(r, 15, data[12])
        sprite_ref = data[6]
        if isinstance(sprite_ref, str) and sprite_ref.startswith("(O)"):
            try:
                oid = int(sprite_ref[3:])
                if oid in GAME_DATA:
                    si.value, pr.value = GAME_DATA[oid]
            except ValueError:
                pass

# ---------- validações ----------
def add_dv(formula, rng):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Valor inválido"; dv.showErrorMessage = True
    ws.add_data_validation(dv); dv.add(rng)

last = first + TOTAL_ROWS - 1
add_dv('"Crop,Forrageável,Animal,Peixe,Monstro,Mineral,Prato,Aldeão,Especial"', f"D{first}:D{last}")
add_dv('"Comum,Incomum,Raro,Lendário"', f"E{first}:E{last}")
add_dv('"Primavera,Verão,Outono,Inverno,Nenhum"', f"F{first}:F{last}")
dv_edge = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
dv_edge.error = "Bordas devem ser de 1 a 10"; dv_edge.showErrorMessage = True
ws.add_data_validation(dv_edge)
for col in "HIJK":
    dv_edge.add(f"{col}{first}:{col}{last}")

# ---------- aba Referências ----------
wr = wb.create_sheet("Referências")
wr["A1"] = "Referências (Stardew Valley Wiki)"; wr["A1"].font = TITLE_FONT
links = [
    ("Wiki (home)", "https://stardewvalleywiki.com/Stardew_Valley_Wiki"),
    ("Villagers",   "https://stardewvalleywiki.com/Villagers"),
    ("Crops",       "https://stardewvalleywiki.com/Crops"),
    ("Animals",     "https://stardewvalleywiki.com/Animals"),
    ("Monsters",    "https://stardewvalleywiki.com/Monsters"),
    ("Foraging",    "https://stardewvalleywiki.com/Foraging"),
    ("Fish",        "https://stardewvalleywiki.com/Fish"),
    ("Cooking",     "https://stardewvalleywiki.com/Cooking"),
    ("Minerals",    "https://stardewvalleywiki.com/Minerals"),
    ("Tools",       "https://stardewvalleywiki.com/Tools"),
]
wr.cell(3, 1, "Tema").font = HDR_FONT; wr.cell(3, 1).fill = HDR_FILL
wr.cell(3, 2, "Link").font = HDR_FONT; wr.cell(3, 2).fill = HDR_FILL
for r, (name, url) in enumerate(links, 4):
    wr.cell(r, 1, name)
    c = wr.cell(r, 2, url); c.hyperlink = url; c.font = Font(color="0563C1", underline="single")
wr.cell(15, 1, "Nota:").font = Font(bold=True)
wr.cell(16, 1, "IDs de objeto VALIDADOS contra o Content/Data/Objects.xnb do jogo instalado (v1.6.15) — "
               "cada (O)<id> confere com o Name interno. Estações das páginas Crops/Foraging.")
wr.column_dimensions["A"].width = 22
wr.column_dimensions["B"].width = 52

wb.save(OUT)
print("salvo:", OUT, "| cartas:", len(seed))
